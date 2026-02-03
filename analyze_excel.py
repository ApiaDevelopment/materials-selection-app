import openpyxl
import json

# Analyze Project1.xlsx
wb1 = openpyxl.load_workbook('Project1.xlsx', data_only=True)
print("=== PROJECT 1 ===")
print(f"Sheets: {[sheet.title for sheet in wb1.worksheets]}\n")

for sheet in wb1.worksheets:
    ws = sheet
    print(f"\n--- Sheet: {ws.title} ---")
    print(f"Dimensions: {ws.dimensions}")
    print("\nFirst 15 rows:")
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        print(f"Row {i}: {row}")

print("\n\n" + "="*80 + "\n\n")

# Analyze Project2.xlsx
wb2 = openpyxl.load_workbook('Project2.xlsx', data_only=True)
print("=== PROJECT 2 ===")
print(f"Sheets: {[sheet.title for sheet in wb2.worksheets]}\n")

for sheet in wb2.worksheets:
    ws = sheet
    print(f"\n--- Sheet: {ws.title} ---")
    print(f"Dimensions: {ws.dimensions}")
    print("\nFirst 15 rows:")
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        print(f"Row {i}: {row}")

print("\n\n" + "="*80 + "\n\n")

# Analyze Material_Selection file
wb3 = openpyxl.load_workbook('Material_Selection_Ordering_and_Received.xlsx', data_only=True)
print("=== MATERIAL SELECTION TEMPLATE ===")
print(f"Sheets: {[sheet.title for sheet in wb3.worksheets]}\n")

for sheet in wb3.worksheets:
    ws = sheet
    print(f"\n--- Sheet: {ws.title} ---")
    print(f"Dimensions: {ws.dimensions}")
    print("\nFirst 15 rows:")
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        print(f"Row {i}: {row}")
